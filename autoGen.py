from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, colors, Color
import string, itertools

letters= string.ascii_uppercase
projects=[]


wb = load_workbook(filename='data.xlsx')

# grab the active worksheet
sheet1 =wb['Sheet1']


#indeces of headers
headers= [x.value.encode('UTF8') for x in sheet1.rows[0]]
# make a headers dictionary
headers_dic = {v:k for k,v in enumerate(headers, start=0)}

project_index = headers.index('PROJECT')

for unit in sheet1.columns[project_index]:
	if unit.value=='PROJECT':
		continue
	if unit.value != None and unit.value not in projects:
		projects.append(unit.value)

projects = [x.encode('UTF8') for x in projects]
projects.sort()


projects_dic = {k:v for (k,v) in enumerate(projects, start=0)}

#dictionary for each table header to new excel column letter
first_row_headers = ['DOC_NO','DOC_SUFFIX','FUND_3','POST_DATE']
letters_dic = {let:header_name for let, header_name in zip(first_row_headers,letters)}


first_row = first_row_headers + projects[2:]+projects[:2]



new_book = Workbook()
new_sheet = new_book.active


new_sheet.append(first_row)
project_col_dic ={}
for col in new_sheet.rows[0]:
	if col.value not in projects:
		continue
	project_col_dic[col.value] = col.column



#project_col_dic = {proj:let for proj,let in zip(projects,project_col)}

ft = Font(color=colors.RED)

for row in sheet1.rows[1:]:
	current_proj = headers_dic['PROJECT']
	current_amt = headers_dic['TRANS_AMT']
	current_doc = headers_dic['DOC_NO']
	current_suffix = headers_dic['DOC_SUFFIX']
	current_fund = headers_dic['FUND_3']
	current_date = headers_dic['POST_DATE']
	current_GL = headers_dic['GL_ACCT']


	# c is for column letter and r is for row number
	c = project_col_dic[row[current_proj].value]
	r = row[current_proj].row
	location = c + str(r)
	new_sheet[letters_dic['DOC_NO'] +str(r)] = row[current_doc].value
	new_sheet[letters_dic['DOC_SUFFIX'] +str(r)] = row[current_suffix].value
	new_sheet[letters_dic['FUND_3'] +str(r)] = row[current_fund].value
	new_sheet[letters_dic['POST_DATE'] +str(r)] = row[current_date].value.date()
	new_sheet[location]= row[current_amt].value
	if row[current_GL].value == '451':
		new_sheet[location].font = ft
	

last_row = new_sheet.max_row

new_sheet.freeze_panes = new_sheet['A2']

for num in projects:
	col = project_col_dic[num]
	place = col + str(last_row+1)
	new_sheet[place]= "=SUM(%s2:%s%d)" %(col,col,last_row)
	new_sheet.column_dimensions[col].width= 13



new_book.save('testing.xlsx')
